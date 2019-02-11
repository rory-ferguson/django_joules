# -*- coding: utf-8 -*-

import requests
import bs4
import logging
import time
from sys import argv
from openpyxl import load_workbook

"""
    Captures the InsecureRequestWarning (HTTPS SSL Verification Error)
"""
logging.captureWarnings(True)

start = time.time()

mega_menu_404, navigation_list_unfiltered, list_purge_duplicates = [], [], []
list_purge_duplicates_one = []
list_purge_duplicates_two = []


class Main(object):

def get_request(url):
    """
    HTTP request to store html as an object
    """
    res = requests.get(url, verify=False)

    if res.status_code == 200:
        global parse_soup
        parse_soup = bs4.BeautifulSoup(res.text, "html.parser")
        # print("The Status Code returned is", res.status_code, url)

    elif res.status_code == 404:
        print("The Status Code returned is", res.status_code, url)
        mega_menu_404.append(url)

    else:
        print("The Status Code returned is", res.status_code, url)


def navigation_unfiltered():
    """
        Stored list of navigation links from the mega menu // unfiltered
    """
    local_list = parse_soup.find_all('li', {'class': 'yCmsComponent mobile-nav-item Lc'})

    for i in local_list:
        navigation_list_unfiltered.append(i.find("a").get("href"))
        # print(i.find("a").get("href"))


def navigation_duplicates(list_arg):
    """
        Function to remove duplicates from a list, preserves original list order
    """

    """ Removes URLs start with &, these are JS injected """
    [list_purge_duplicates_one.append(i) for i in list_arg if not i.startswith('&')]

    """ Removes tracking """
    for i in list_purge_duplicates_one:
        if '?' in i:
            x = i.split('?')[0]
            list_purge_duplicates_two.append(x)
        elif '?' not in i:
            list_purge_duplicates_two.append(i)

    """ Removes duplicate URLS """
    [list_purge_duplicates.append(i) for i in list_purge_duplicates_two if not list_purge_duplicates.count(i)]

    return list_purge_duplicates


def remove_external_urls():
    """
        Filter out external urls such as blog/pinterest
    """
    global purge_external_urls
    purge_external_urls = list(list_purge_duplicates)

    [purge_external_urls.remove(i) for i in list_purge_duplicates if i.__contains__('.')]


def category(domain, url, name):
    wb = load_workbook('Template.xlsx')
    ws = wb.active
    counter = 0

    for title in url:
        for i in range(100):
            html_doc = requests.get(str(domain) + str(title) + "?showFragment=true&page=" + str(i),
                                    verify=False,
                                    allow_redirects=False)
            soup = bs4.BeautifulSoup(html_doc.text, "html.parser")
            tags = soup.find_all("div", {"class": "product-grid-item-inner"})
            if not len(soup) == 0:
                for i in tags:
                    """ New Price """
                    new_price = None
                    try:
                        new_temp = i.find('div', {'class': 'product-price'})
                        new_price = new_temp.find('span', class_='new-price').text
                    except:
                        pass

                    """ Was Price """
                    was_price = None
                    try:
                        c = new_temp.find("span", class_='old-price')
                        was_temp = c.find("div", {"id": "was-"}).text
                        was_price = was_temp.strip(' \t\n\r')
                    except:
                        pass
                    try:
                        was_temp = c.find("div", {"id": "wasFormated-"}).text
                        was_price = was_temp.strip(' \t\n\r')
                    except:
                        pass

                    """ Was Was Price """
                    waswas_price = None
                    try:
                        waswas_temp = c.find("div", {"id": "wasWas-"}).text
                        waswas_price = waswas_temp.strip(' \t\n\r')
                    except:
                        pass
                    try:
                        waswas_temp = c.find("div", {"id": "wasWasFormated-"}).text
                        waswas_price = waswas_temp.strip(' \t\n\r')
                    except:
                        pass

                    """ Image Badge Name """
                    img_badge = None
                    try:
                        o = i.find('div', {'id': 'badges'})
                        img_string = o.find('img')['src']
                        f = img_string.split('medias/')[1]
                        img_badge = f.split('?')[0]
                    except:
                        pass

                    """ Product ID """
                    product_id = None
                    product_href = None
                    try:
                        r = i.find('div', {'class': 'product-name'})
                        product_href = r.a['href']
                        product_id = product_href.split('?id=')[1]
                    except:
                        pass

                    """
                        get request to product landing page,
                        to collect 404, prices etc
                    """
                    product_list = []
                    product_list.extend((html_doc.status_code,
                                        title,
                                        product_id,
                                        img_badge,
                                        waswas_price,
                                        was_price,
                                        new_price))
                    print(product_list)

                    counter += 1
                    for col, val in enumerate(product_list, start=1):
                        ws.cell(row=counter + 2, column=col).value = val
            else:
                break

    wb.save(filename=f'{name}.xlsx')


def main():
    """this is the main file that the script is run from"""

    domain_url = 'https://www.joules.com'
    print(f'\nUsing {domain_url} as the domain url\n')

    get_request(url=domain_url)

    navigation_unfiltered()

    navigation_duplicates(list_arg=navigation_list_unfiltered)

    remove_external_urls()

    category(domain=domain_url, url=purge_external_urls, name=f'main')

    end = time.time()
    print(f'\nThe script took {end - start} seconds to run')


main()
